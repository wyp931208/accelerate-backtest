# -*- coding: utf-8 -*-
"""
XPK加速策略 - Streamlit回测与信号提醒系统
主应用入口
"""
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from io import BytesIO
import sys
import os

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from data_service import (
    get_pro_api, get_latest_trade_date, get_trade_calendar,
    get_stock_basic, get_daily_data_with_info,
    get_stock_kline_data, get_stock_weekly_kline, DEFAULT_TOKEN
)
from backtest_engine import run_backtest, detect_daily_signals
from charts import plot_daily_kline_with_indicators, plot_weekly_kline_with_indicators
from pdf_export import export_charts_to_pdf, export_multi_stocks_pdf

# ======================
# 页面配置
# ======================
st.set_page_config(
    page_title="XPK加速策略",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ======================
# 初始化Session State
# ======================
if 'tushare_token' not in st.session_state:
    st.session_state.tushare_token = DEFAULT_TOKEN
if 'backtest_results' not in st.session_state:
    st.session_state.backtest_results = None

latest_trade_date_str = get_latest_trade_date()
latest_trade_date_dt = datetime.strptime(latest_trade_date_str, '%Y%m%d').date()

# ======================
# 侧边栏 - Token配置
# ======================
with st.sidebar:
    st.markdown("### 🔑 Tushare配置")
    token_input = st.text_input(
        "Tushare API Token",
        value=st.session_state.tushare_token,
        type="password",
        help="已内置默认Token，可直接使用。也可替换为你自己的Token"
    )
    if token_input:
        st.session_state.tushare_token = token_input

    st.success("Token已配置 ✓")

    st.markdown("---")

# ======================
# 工具函数
# ======================
def to_excel(df: pd.DataFrame) -> bytes:
    """将DataFrame导出为Excel字节流（单Sheet）"""
    from openpyxl.utils import get_column_letter
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        _auto_fit_columns(writer.sheets['Sheet1'], df)
    output.seek(0)
    return output.getvalue()


def to_excel_multi_sheets(sheets: dict) -> bytes:
    """
    将多个DataFrame导出为Excel字节流（多Sheet）
    sheets: dict, {sheet_name: DataFrame}
    """
    from openpyxl.utils import get_column_letter
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            _auto_fit_columns(writer.sheets[sheet_name], df)
    output.seek(0)
    return output.getvalue()


def _auto_fit_columns(worksheet, df: pd.DataFrame):
    """自动调整Excel列宽"""
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(df.columns):
        col_values = df[col].to_numpy(dtype=object, na_value='')
        str_lens = [len(str(v)) for v in col_values]
        max_len = max(max(str_lens, default=0), len(str(col)))
        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 30)

# ======================
# 主导航
# ======================
tab1, tab2, tab3 = st.tabs(["📊 策略回测", "🔔 信号提醒", "📉 K线图表"])

# ============================================================
# TAB 1: 策略回测
# ============================================================
with tab1:
    st.header("📊 XPK加速策略回测")

    # 参数设置区域
    with st.expander("⚙️ 回测参数设置", expanded=True):
        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("**基本信息**")
            start_date = st.date_input(
                "开始日期",
                value=min(datetime(2025, 1, 1).date(), latest_trade_date_dt),
                max_value=latest_trade_date_dt,
                key="bt_start"
            )
            end_date = st.date_input(
                "结束日期",
                value=latest_trade_date_dt,
                max_value=latest_trade_date_dt,
                key="bt_end"
            )
            adj_type = st.selectbox(
                "复权方式",
                options=["qfq", "hfq", "none"],
                format_func=lambda x: {"qfq": "前复权", "hfq": "后复权", "none": "不复权"}[x],
                index=0,
                key="bt_adj_type",
                help="前复权：以最新价格为基准向前调整（推荐）；后复权：以上市首日为基准向后调整；不复权：原始价格"
            )
            buy_amount = st.number_input(
                "单次买入金额(元)", min_value=10000, max_value=1000000,
                value=100000, step=10000, key="bt_buy_amount"
            )
            max_hold_days = st.number_input(
                "最大持有天数", min_value=5, max_value=60,
                value=20, key="bt_max_hold"
            )
            supplement_rate = st.number_input(
                "补仓比例(跌破成本的%)", min_value=0.70, max_value=0.95,
                value=0.86, step=0.01, format="%.2f", key="bt_supp_rate"
            )

        with col2:
            st.markdown("**信号筛选条件**")
            volume_ratio_min = st.number_input(
                "量比下限", min_value=1.0, max_value=5.0,
                value=1.5, step=0.1, key="bt_vr_min"
            )
            volume_ratio_max = st.number_input(
                "量比上限", min_value=1.5, max_value=10.0,
                value=2.5, step=0.1, key="bt_vr_max"
            )
            pct_chg_min = st.number_input(
                "涨跌幅下限(%)", min_value=0.0, max_value=10.0,
                value=2.0, step=0.5, key="bt_pct_min"
            )
            pct_chg_max = st.number_input(
                "涨跌幅上限(%)", min_value=2.0, max_value=20.0,
                value=8.0, step=0.5, key="bt_pct_max"
            )

        with col3:
            st.markdown("**上影线与涨幅限制**")
            upper_shadow_min = st.number_input(
                "上影线比例下限", min_value=0.0, max_value=0.8,
                value=0.25, step=0.05, format="%.2f", key="bt_us_min"
            )
            upper_shadow_max = st.number_input(
                "上影线比例上限", min_value=0.1, max_value=1.0,
                value=0.50, step=0.05, format="%.2f", key="bt_us_max"
            )
            require_cum_pct = st.checkbox("要求前N日累计涨幅", value=True, key="bt_req_cum_pct",
                                          help="启用后将筛选前N日累计涨幅在指定范围内的股票；关闭则不限制累计涨幅")
            n_days_lookback = st.number_input(
                "前N日累计涨幅(N)", min_value=3, max_value=60,
                value=20, key="bt_n_days",
                disabled=not require_cum_pct
            )
            cum_pct_min = st.number_input(
                "?????????(%)", min_value=-100.0, max_value=100.0,
                value=0.0, step=5.0, key="bt_cum_min",
                disabled=not require_cum_pct
            )
            cum_pct_max = st.number_input(
                "?????????(%)", min_value=-100.0, max_value=100.0,
                value=20.0, step=5.0, key="bt_cum_max",
                disabled=not require_cum_pct
            )

            require_vwap = st.checkbox("要求收盘高于VWAP", value=True, key="bt_req_vwap")
            vwap_pct = st.number_input(
                "收盘高于VWAP百分比(%)", min_value=0.0, max_value=2.0,
                value=0.3, step=0.1, key="bt_vwap_pct",
                disabled=not require_vwap
            )

    # 回测执行
    st.info("💡 **提示**：如果回测显示「未产生有效交易」，请尝试：\n"
            "1. 降低量比下限（如1.0）或扩大量比上限（如5.0）\n"
            "2. 降低累计涨幅下限（如10%）\n"
            "3. 降低上影线比例下限（如0.10）\n"
            "4. 取消「要求收盘高于VWAP」\n"
            "5. 扩大日期范围以覆盖更多交易日\n\n"
            "📌 **日期说明**：为计算前N日累计涨幅，数据获取会自动从起始日期往前扩展N个交易日；"
            "为保证持股周期完整，有效信号截止日会自动从结束日期往前缩减最大持股天数个交易日。")
    if st.button("🚀 开始回测", type="primary", use_container_width=True):
        params = {
            "buy_amount": buy_amount,
            "lot_size": 100,
            "max_hold_days": max_hold_days,
            "supplement_rate": supplement_rate,
            "volume_ratio_min": volume_ratio_min,
            "volume_ratio_max": volume_ratio_max,
            "pct_chg_min": pct_chg_min,
            "pct_chg_max": pct_chg_max,
            "upper_shadow_ratio_min": upper_shadow_min,
            "upper_shadow_ratio_max": upper_shadow_max,
            "require_close_above_vwap": require_vwap,
            "close_above_vwap_pct": vwap_pct,
            "n_days_lookback": n_days_lookback,
            "cum_pct_chg_min": cum_pct_min,
            "cum_pct_chg_max": cum_pct_max,
            "require_cum_pct": require_cum_pct,
            "adj_type": adj_type,
        }

        # ========== 计算数据获取的实际日期范围 ==========
        # 1. 数据起始日期：往前扩展 n_days_lookback + 10 个交易日，确保累计涨幅有足够历史
        # 2. 有效回测截止日期：从结束日期往前缩减 max_hold_days 个交易日，保证持股周期完整
        cal_df = get_trade_calendar("20200101", end_date.strftime('%Y%m%d'))
        if not cal_df.empty:
            open_dates = cal_df[cal_df['is_open'] == 1]['cal_date'].sort_values().tolist()
        else:
            open_dates = []

        # 计算数据获取起始日期：从用户选择的起始日期往前找 N+10 个交易日
        sd_str = start_date.strftime('%Y%m%d')
        if open_dates:
            # 找到起始日期之前 N+10 个交易日
            dates_before_start = [d for d in open_dates if d <= sd_str]
            extend_count = n_days_lookback + 10
            if len(dates_before_start) > extend_count:
                data_start_date = dates_before_start[-extend_count - 1] if len(dates_before_start) > extend_count + 1 else dates_before_start[0]
                # 更精确：取起始日期往前第 extend_count 个交易日
                start_idx = len(dates_before_start) - 1  # 起始日期在列表中的位置
                data_start_idx = max(0, start_idx - extend_count)
                data_start_date = dates_before_start[data_start_idx]
            else:
                data_start_date = dates_before_start[0] if dates_before_start else sd_str
        else:
            data_start_date = sd_str

        # 计算有效回测截止日期：从结束日期往前找 max_hold_days 个交易日
        ed_str = end_date.strftime('%Y%m%d')
        if open_dates:
            dates_before_end = [d for d in open_dates if d <= ed_str]
            if len(dates_before_end) > max_hold_days:
                effective_end_idx = len(dates_before_end) - 1 - max_hold_days
                effective_end_date = dates_before_end[max(0, effective_end_idx)]
            else:
                effective_end_date = dates_before_end[0] if dates_before_end else ed_str
        else:
            effective_end_date = ed_str

        # 传入回测引擎的有效信号起止日期
        params["start_date"] = sd_str
        params["end_date"] = effective_end_date

        with st.spinner("正在获取数据并执行回测，请耐心等待..."):
            # 数据获取范围：从扩展后的起始日期到用户选择的结束日期
            daily = get_daily_data_with_info(data_start_date, ed_str, adj_type=adj_type)
            if daily.empty:
                st.error("获取数据失败，请检查Token和网络连接")
            else:
                st.info(f"获取到 {len(daily)} 条日线数据\n\n"
                        f"📊 **数据范围**：{data_start_date} ~ {ed_str}（已向前扩展{n_days_lookback + 10}个交易日用于累计涨幅计算）\n"
                        f"📈 **有效回测区间**：{sd_str} ~ {effective_end_date}（已排除最后{max_hold_days}个交易日以保证持股周期完整）")

                def progress_cb(progress, text):
                    st.progress(progress, text=text)

                df_summary, df_all_trades, signals = run_backtest(
                    daily, params, progress_callback=progress_cb
                )

                st.session_state.backtest_results = {
                    'summary': df_summary,
                    'all_trades': df_all_trades,
                    'signals': signals
                }

    # 展示回测结果
    if st.session_state.backtest_results:
        results = st.session_state.backtest_results
        df_summary = results['summary']
        df_all_trades = results['all_trades']
        signals = results['signals']

        if df_summary.empty:
            st.warning("未产生有效交易，请调整筛选条件")
        else:
            st.markdown("### 📈 回测结果汇总")

            # 最佳配置
            best_row = df_summary.loc[df_summary["单位时间收益率"].idxmax()]
            best_total_row = df_summary.loc[df_summary["整体收益率"].idxmax()]

            col_best1, col_best2 = st.columns(2)
            with col_best1:
                st.metric(
                    "🏆 资金效率最高",
                    f"{best_row['板块']} 止盈{best_row['止盈点(%)']}%",
                    f"单位时间收益率: {best_row['单位时间收益率']:.4f}"
                )
            with col_best2:
                st.metric(
                    "💰 总收益最高",
                    f"{best_total_row['板块']} 止盈{best_total_row['止盈点(%)']}%",
                    f"整体收益率: {best_total_row['整体收益率']:.4f}"
                )

            # 按板块展示汇总
            for board in df_summary["板块"].unique():
                st.markdown(f"#### {board}")
                board_summary = df_summary[df_summary["板块"] == board]
                st.dataframe(
                    board_summary.style.format({
                        "胜率": "{:.2%}",
                        "单位时间收益率": "{:.4f}",
                        "整体收益率": "{:.4f}",
                        "总投入": "{:,.0f}",
                        "总盈利": "{:,.0f}",
                    }),
                    use_container_width=True,
                    hide_index=True
                )

            # ── 交易明细大表 ──
            if not df_all_trades.empty:
                st.markdown("### 📋 完整交易明细")
                st.markdown(f"共 **{len(df_all_trades)}** 条交易记录")

                # 筛选控件
                filter_col1, filter_col2, filter_col3 = st.columns(3)
                with filter_col1:
                    pt_options = sorted(df_all_trades["止盈目标(%)"].unique())
                    selected_pt = st.multiselect(
                        "止盈目标(%)", pt_options, default=pt_options, key="bt_detail_pt"
                    )
                with filter_col2:
                    board_options = list(df_all_trades["板块"].unique())
                    selected_boards = st.multiselect(
                        "板块", board_options, default=board_options, key="bt_detail_board"
                    )
                with filter_col3:
                    sell_type_options = sorted(df_all_trades["卖出方式"].unique())
                    selected_sell_types = st.multiselect(
                        "卖出方式", sell_type_options, default=sell_type_options, key="bt_detail_sell_type"
                    )

                # 应用筛选
                detail_df = df_all_trades[
                    df_all_trades["止盈目标(%)"].isin(selected_pt) &
                    df_all_trades["板块"].isin(selected_boards) &
                    df_all_trades["卖出方式"].isin(selected_sell_types)
                ].copy()

                if not detail_df.empty:
                    # 汇总统计卡片
                    stat_col1, stat_col2, stat_col3, stat_col4, stat_col5 = st.columns(5)
                    with stat_col1:
                        st.metric("交易数", f"{len(detail_df)}")
                    with stat_col2:
                        win_count = len(detail_df[detail_df["盈利金额"] > 0])
                        st.metric("盈利次数", f"{win_count}")
                    with stat_col3:
                        wr = win_count / len(detail_df) if len(detail_df) > 0 else 0
                        st.metric("胜率", f"{wr:.2%}")
                    with stat_col4:
                        total_profit = detail_df["盈利金额"].sum()
                        st.metric("总盈利", f"{total_profit:,.0f}")
                    with stat_col5:
                        avg_ret = detail_df["收益率"].mean()
                        st.metric("平均收益率", f"{avg_ret:.2%}")

                    # 交易明细表格
                    display_df = detail_df.copy()
                    # 将None/NaT替换为空字符串，避免style.format报错
                    for col in ["补仓日期", "补仓价格", "补仓股数"]:
                        if col in display_df.columns:
                            display_df[col] = display_df[col].fillna("").astype(object)
                            display_df[col] = display_df[col].replace({pd.NaT: ""})

                    format_dict = {}
                    for col, fmt in [
                        ("收益率", "{:.2%}"),
                        ("初始买入价", "{:.3f}"),
                        ("卖出价格", "{:.3f}"),
                        ("补仓后成本", "{:.2f}"),
                        ("投入金额", "{:,.2f}"),
                        ("盈利金额", "{:,.2f}"),
                    ]:
                        if col in display_df.columns:
                            format_dict[col] = fmt
                    # 补仓价格单独处理：空字符串不格式化
                    if "补仓价格" in display_df.columns:
                        format_dict["补仓价格"] = lambda x: f"{x:.3f}" if x != "" and not pd.isna(x) else ""

                    st.dataframe(
                        display_df.style.format(format_dict, na_rep=""),
                        use_container_width=True,
                        hide_index=True,
                        height=min(600, max(200, len(display_df) * 35 + 50))
                    )
                else:
                    st.info("当前筛选条件下无交易记录")

            # ── 下载区域 ──
            st.markdown("### 📥 下载回测结果")
            dl_col1, dl_col2, dl_col3 = st.columns(3)
            with dl_col1:
                summary_xlsx = to_excel(df_summary)
                st.download_button(
                    "📥 下载回测汇总 (Excel)",
                    data=summary_xlsx,
                    file_name="回测结果汇总.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with dl_col2:
                if not df_all_trades.empty:
                    trades_xlsx = to_excel(df_all_trades)
                    st.download_button(
                        "📥 下载全部交易明细 (Excel)",
                        data=trades_xlsx,
                        file_name="全部交易明细.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with dl_col3:
                if not df_all_trades.empty:
                    # 综合Excel：多Sheet包含汇总+各止盈点明细+全部明细
                    sheets = {"回测汇总": df_summary}
                    for pt in sorted(df_all_trades["止盈目标(%)"].unique()):
                        pt_df = df_all_trades[df_all_trades["止盈目标(%)"] == pt]
                        sheet_name = f"止盈{pt}%"
                        # Excel sheet名称最长31字符
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:28] + "..."
                        sheets[sheet_name] = pt_df.reset_index(drop=True)
                    sheets["全部交易明细"] = df_all_trades.reset_index(drop=True)
                    full_xlsx = to_excel_multi_sheets(sheets)
                    st.download_button(
                        "📥 下载综合报告 (多Sheet Excel)",
                        data=full_xlsx,
                        file_name="XPK加速策略回测报告.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )


# ============================================================
# TAB 2: 信号提醒
# ============================================================
with tab2:
    st.header("🔔 前一日信号提醒")

    st.markdown("""
    基于Tushare日线数据，检测前一交易日的加速策略信号。
    由于Tushare日线数据为当日收盘后更新，因此展示的是**前一个交易日**的信号。
    """)

    # 信号检测参数（与回测共享默认值）
    with st.expander("⚙️ 信号参数设置", expanded=False):
        col_s1, col_s2, col_s3 = st.columns(3)
        with col_s1:
            st.markdown("**量比与涨跌幅**")
            sig_vr_min = st.number_input("量比下限", value=1.5, step=0.1, key="sig_vr_min")
            sig_vr_max = st.number_input("量比上限", value=2.5, step=0.1, key="sig_vr_max")
            sig_pct_min = st.number_input("涨跌幅下限(%)", value=2.0, step=0.5, key="sig_pct_min")
            sig_pct_max = st.number_input("涨跌幅上限(%)", value=8.0, step=0.5, key="sig_pct_max")
        with col_s2:
            st.markdown("**上影线与累计涨幅**")
            sig_us_min = st.number_input("上影线下限", value=0.25, step=0.05, format="%.2f", key="sig_us_min")
            sig_us_max = st.number_input("上影线上限", value=0.50, step=0.05, format="%.2f", key="sig_us_max")
            sig_req_cum_pct = st.checkbox("要求前N日累计涨幅", value=True, key="sig_req_cum_pct",
                                          help="启用后将筛选前N日累计涨幅在指定范围内的股票")
            sig_cum_min = st.number_input("?????????(%)", min_value=-100.0, max_value=100.0, value=0.0, step=5.0, key="sig_cum_min",
                                         disabled=not sig_req_cum_pct)
            sig_cum_max = st.number_input("?????????(%)", min_value=-100.0, max_value=100.0, value=20.0, step=5.0, key="sig_cum_max",
                                         disabled=not sig_req_cum_pct)
        with col_s3:
            st.markdown("**累计涨幅天数与VWAP**")
            sig_n_days = st.number_input("前N日累计涨幅(N)", min_value=3, max_value=60, value=20, step=1, key="sig_n_days",
                                         disabled=not sig_req_cum_pct)
            sig_req_vwap = st.checkbox("要求收盘高于VWAP", value=True, key="sig_req_vwap")
            sig_vwap_pct = st.number_input(
                "收盘高于VWAP百分比(%)", min_value=0.0, max_value=2.0,
                value=0.3, step=0.1, key="sig_vwap_pct",
                disabled=not sig_req_vwap
    # ????????????
            )
    latest_td = latest_trade_date_str
    st.info(f"?????????: **{latest_td}**")

    col_date1, col_date2 = st.columns([3, 1])
    with col_date1:
        signal_date_dt = st.date_input(
            "????",
            value=latest_trade_date_dt,
            max_value=latest_trade_date_dt,
            key="sig_date"
        )
        signal_date = signal_date_dt.strftime("%Y%m%d")
    with col_date2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("?? ??????"):
            st.session_state.sig_date = latest_trade_date_dt
            st.rerun()

    if st.button("🔍 检测信号", type="primary", use_container_width=True):
        sig_params = {
            "volume_ratio_min": sig_vr_min,
            "volume_ratio_max": sig_vr_max,
            "pct_chg_min": sig_pct_min,
            "pct_chg_max": sig_pct_max,
            "upper_shadow_ratio_min": sig_us_min,
            "upper_shadow_ratio_max": sig_us_max,
            "cum_pct_chg_min": sig_cum_min,
            "cum_pct_chg_max": sig_cum_max,
            "n_days_lookback": sig_n_days,
            "require_cum_pct": sig_req_cum_pct,
            "require_close_above_vwap": sig_req_vwap,
            "close_above_vwap_pct": sig_vwap_pct,
        }

        with st.spinner("正在检测信号，需要逐股票查询历史数据，请耐心等待..."):
            signal_df = detect_daily_signals(signal_date, sig_params)

        if signal_df.empty:
            st.warning(f"日期 {signal_date} 未检测到加速策略信号")
        else:
            st.success(f"检测到 {len(signal_df)} 条信号！")

            st.dataframe(
                signal_df.style.format({
                    "涨跌幅(%)": "{:.2f}",
                    "量比": "{:.2f}",
                    "上影线比例": "{:.4f}",
                    "前N日累计涨幅(%)": "{:.2f}",
                    "收盘价": "{:.2f}",
                    "VWAP": "{:.2f}",
                }),
                use_container_width=True,
                hide_index=True
            )

            # 下载信号
            sig_xlsx = to_excel(signal_df)
            st.download_button(
                "📥 下载信号列表 (Excel)",
                data=sig_xlsx,
                file_name=f"加速策略信号_{signal_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # 存储信号供K线图页面使用
            st.session_state.latest_signals = signal_df
            st.session_state.signal_date = signal_date


# ============================================================
# TAB 3: K线图表
# ============================================================
with tab3:
    st.header("📉 K线图表与PDF下载")

    st.markdown("""
    查看信号股票近40个交易日的K线图、成交量、MACD、PSY指标，
    支持日K和周K，可筛选后下载PDF报告。
    """)

    # 输入方式
    input_mode = st.radio(
        "选择股票方式",
        ["手动输入", "从信号提醒获取"],
        horizontal=True,
        key="chart_input_mode"
    )

    selected_codes = []

    if input_mode == "手动输入":
        code_input = st.text_input(
            "输入股票代码（多个用逗号分隔，如 300001.SZ,300002.SZ）",
            key="chart_codes"
        )
        if code_input:
            selected_codes = [c.strip() for c in code_input.split(",") if c.strip()]
    else:
        if 'latest_signals' in st.session_state and not st.session_state.latest_signals.empty:
            sig_df = st.session_state.latest_signals
            signal_stocks = sig_df[['ts_code', '股票名称']].values.tolist()
            options = [f"{code} - {name}" for code, name in signal_stocks]
            selected_options = st.multiselect(
                "选择要查看的信号股票",
                options,
                key="chart_signal_select"
            )
            selected_codes = [opt.split(" - ")[0] for opt in selected_options]
        else:
            st.info("请先在「信号提醒」页面检测信号")

    end_date_chart_dt = st.date_input(
        "????",
        value=latest_trade_date_dt,
        max_value=latest_trade_date_dt,
        key="chart_end_date"
    )
    end_date_chart = end_date_chart_dt.strftime("%Y%m%d")

    n_days_chart = st.slider("??????", min_value=20, max_value=80, value=40,
                              key="chart_n_days")

    if selected_codes and st.button("?? ????", type="primary", use_container_width=True):
        chart_results = []

        for ts_code in selected_codes:
            st.markdown(f"---\n### {ts_code}")

            with st.spinner(f"???? {ts_code} ????????..."):
                stock_info_df = get_stock_basic()
                name_row = stock_info_df[stock_info_df["ts_code"] == ts_code]
                stock_name = name_row["name"].values[0] if not name_row.empty else ts_code

                daily_df = get_stock_kline_data(ts_code, end_date_chart, n_days_chart)
                if daily_df.empty:
                    st.warning(f"{ts_code} ????????")
                    continue

                daily_buf = plot_daily_kline_with_indicators(daily_df, ts_code, stock_name)
                st.image(daily_buf, caption=f"{stock_name} ???", use_container_width=True)

                weekly_df = get_stock_weekly_kline(ts_code, end_date_chart, n_weeks=40)
                weekly_buf = None
                if not weekly_df.empty:
                    weekly_buf = plot_weekly_kline_with_indicators(weekly_df, ts_code, stock_name)
                    st.image(weekly_buf, caption=f"{stock_name} ???", use_container_width=True)

                chart_results.append({
                    "ts_code": ts_code,
                    "name": stock_name,
                    "daily_buf": daily_buf,
                    "weekly_buf": weekly_buf,
                    "stock_info": {
                        "ts_code": ts_code,
                        "name": stock_name,
                        "signal_date": st.session_state.get("signal_date", end_date_chart),
                        "board": "???" if ts_code.startswith("300") else "??",
                    }
                })

        st.session_state.chart_results = chart_results

    # PDF下载区域
    if 'chart_results' in st.session_state and st.session_state.chart_results:
        st.markdown("---")
        st.markdown("### 📥 PDF下载")

        chart_results = st.session_state.chart_results

        # 选择要下载的股票
        pdf_options = [f"{r['ts_code']} - {r['name']}" for r in chart_results]
        selected_pdf = st.multiselect(
            "选择要导出PDF的股票",
            pdf_options,
            default=pdf_options,
            key="pdf_select"
        )

        if selected_pdf and st.button("📄 生成PDF报告"):
            with st.spinner("正在生成PDF..."):
                stocks_data = []
                for opt in selected_pdf:
                    ts_code = opt.split(" - ")[0]
                    for r in chart_results:
                        if r['ts_code'] == ts_code:
                            buffers = [r['daily_buf']]
                            if r['weekly_buf']:
                                buffers.append(r['weekly_buf'])
                            stocks_data.append({
                                'chart_buffers': buffers,
                                'stock_info': r['stock_info']
                            })

                if len(stocks_data) == 1:
                    # 单只股票
                    pdf_buf = export_charts_to_pdf(
                        stocks_data[0]['chart_buffers'],
                        stocks_data[0]['stock_info']
                    )
                else:
                    # 多只股票
                    pdf_buf = export_multi_stocks_pdf(stocks_data)

                st.download_button(
                    "📥 下载PDF报告",
                    data=pdf_buf.getvalue(),
                    file_name=f"加速策略分析报告_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf"
                )


# ======================
# 页脚
# ======================
st.markdown("---")
footer_col1, footer_col2 = st.columns([3, 1])
with footer_col1:
    st.markdown(
        "<div style='text-align: center; color: gray; font-size: 12px;'>"
        "XPK加速策略回测系统 | 数据来源: Tushare Pro | "
        "仅供学习研究，不构成投资建议"
        "</div>",
        unsafe_allow_html=True
    )
with footer_col2:
    # 用户手册下载
    manual_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "user_manual.md")
    if os.path.exists(manual_path):
        with open(manual_path, "r", encoding="utf-8") as f:
            manual_content = f.read()
        st.download_button(
            "📖 下载用户手册",
            data=manual_content,
            file_name="XPK加速策略用户手册.md",
            mime="text/markdown"
        )
